import openpyxl
import os

# Указываем путь к файлу
print(f"Текущая директория: {os.getcwd()}")  # Вывод текущей директории
FILE_NAME = os.path.join(os.getcwd(), "Management/acc_reg_data.xlsx")

# Создание Excel-файла
def create_excel_file():
    print("Функция create_excel_file вызвана.")  # Проверка
    # Создаём папку Management, если её нет
    directory = os.path.dirname(FILE_NAME)
    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Создана директория {directory}")

    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"

        # Заголовки столбцов по умолчанию
        headers = ["ID", "profile", "name", "reg_ip", "tel", "pass", "email"]
        sheet.append(headers)

        # Сохраняем файл
        workbook.save(FILE_NAME)
        print(f"Файл {FILE_NAME} успешно создан!")
    else:
        print(f"Файл {FILE_NAME} уже существует.")

# Добавление записи в таблицу
def add_record_to_excel():
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active

    # Получаем текущие заголовки столбцов
    headers = [cell.value for cell in sheet[1]]

    # Подготовка данных
    data = []
    for header in headers:
        if header == "ID":  # ID заполняется автоматически
            continue
        value = input(f"Введите {header}: ")
        data.append(value)

    # Определяем следующий ID (номер строки)
    next_id = sheet.max_row

    # Добавляем запись с новым ID
    record = [next_id] + data
    sheet.append(record)

    # Сохраняем изменения
    workbook.save(FILE_NAME)
    print("Запись успешно добавлена!")

# Открытие файла Excel
def open_excel_file():
    try:
        os.startfile(FILE_NAME)
        print(f"Файл {FILE_NAME} открыт в Excel.")
    except Exception as e:
        print(f"Не удалось открыть файл: {e}")

# Настройка столбцов
def configure_columns():
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active

    # Считываем текущие заголовки и строки
    headers = [cell.value for cell in sheet[1]]  # Заголовки
    data = [
        [cell.value for cell in row] for row in sheet.iter_rows(min_row=2)
    ]  # Данные строк

    # Меню конфигурации
    print("\nТекущие столбцы:", headers)
    print("1. Добавить столбец")
    print("2. Удалить столбец")
    print("3. Переименовать столбец")
    choice = input("Выберите действие: ")

    if choice == "1":  # Добавить столбец
        new_column = input("Введите название нового столбца: ")
        headers.append(new_column)
        for row in data:
            row.append("")  # Добавляем пустое значение для новых столбцов
    elif choice == "2":  # Удалить столбец
        delete_column = input("Введите название столбца для удаления: ")
        if delete_column in headers:
            index = headers.index(delete_column)  # Индекс удаляемого столбца
            headers.pop(index)  # Удаляем заголовок
            for row in data:
                row.pop(index)  # Удаляем данные соответствующего столбца
            print(f"Столбец {delete_column} успешно удалён!")
        else:
            print("Такого столбца нет.")
            return
    elif choice == "3":  # Переименовать столбец
        old_column = input("Введите название столбца для переименования: ")
        if old_column in headers:
            new_name = input("Введите новое название столбца: ")
            index = headers.index(old_column)
            headers[index] = new_name
            print(f"Столбец {old_column} переименован в {new_name}.")
        else:
            print("Такого столбца нет.")
            return
    else:
        print("Неверный выбор.")
        return

    # Очистка и перезапись таблицы с обновлёнными данными
    sheet.delete_rows(1, sheet.max_row)  # Удаляем все данные
    sheet.append(headers)  # Добавляем новые заголовки
    for row in data:
        sheet.append(row)  # Добавляем обновлённые строки

    # Сохраняем изменения
    workbook.save(FILE_NAME)
    print("Изменения успешно сохранены!")



# Основное меню
def main():
    print("Запуск main()")  # Отладочный вывод
    create_excel_file()  # Создаём файл, если его нет
    while True:
        print("\nМеню:")
        print("1. Добавить запись в таблицу")
        print("2. Открыть файл в Excel")
        print("3. Настроить столбцы")
        print("4. Выйти")

        choice = input("Выберите действие: ")
        print(f"Введено значение: {choice}")  # Отладочный вывод
        if choice == "1":
            add_record_to_excel()
        elif choice == "2":
            open_excel_file()
        elif choice == "3":
            configure_columns()
        elif choice == "4":
            print("Выход из программы.")
            break
        else:
            print("Неверный выбор. Попробуйте снова.")

if __name__ == "__main__":
    main()
